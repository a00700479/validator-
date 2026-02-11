import argparse
import csv
import os
import re
import subprocess
import sys
import itertools
import tempfile
import ipaddress
import io
from dataclasses import dataclass
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from collections import defaultdict
from typing import Iterator

from services.report_service import ReportService
from config.columns import (
    COLUMN_NAMES,
    REQUIRED_COLUMNS,
    VALID_PROTOCOLS,
    VALID_CHECK_METHODS,
)

print("=== PRECHECK STARTED ===", flush=True)

BASE_DIR = Path(__file__).resolve().parent

# Какие разделители пытаемся распознать
DELIMS = ["|", ";", ",", "\t"]

MASK_RE = re.compile(r"/\s*(\d{1,3})")


# ----------------------------
# Чтение файла с возможно разными кодировками (вдруг не utf-8)
# --------------------------


def detect_encoding_and_text(path: Path) -> tuple[str, str]:
    """
    Пытаемся прочитать файл в разных кодировках.
    Возвращаем (кодировка, текст).
    """
    raw = path.read_bytes()

    # --- 1) ЯВНЫЙ  BOM (самое важное) ---
    if raw.startswith(b"\xff\xfe"):
        return "utf-16-le", raw.decode("utf-16-le", errors="replace")
    if raw.startswith(b"\xfe\xff"):
        return "utf-16-be", raw.decode("utf-16-be", errors="replace")
    if raw.startswith(b"\xef\xbb\xbf"):
        return "utf-8", raw.decode("utf-8-sig", errors="replace")

    # Сначала пробуем норму по ТЗ: UTF-8
    for enc in ("utf-8",):
        try:
            return enc, raw.decode(enc)
        except UnicodeDecodeError:
            pass
    # Дополнимтельно
    for enc in ("utf-16",):
        try:
            return enc, raw.decode(enc, errors="replace")
        except Exception:
            pass
    # Фолбэки (в реальности часто встречаются)
    for enc in ("cp1251", "cp866", "latin-1"):
        try:
            return enc, raw.decode(enc, errors="replace")
        except Exception:
            continue
    # Последний шанс - не упасть вообще
    return "unknown", raw.decode("utf-8", errors="replace")


# -------------------------------------------------
# Работа с разделителем и заголовками
# -------------------------------------------------
def detect_delimiter(header_line: str) -> tuple[str | None, dict]:
    """
    Определяем разделитель по строке заголовков.
    Всегда возвращаем (delim, counts).
    Если разделитель не найден - delim = None
    """
    counts = {d: header_line.count(d) for d in DELIMS}
    best = max(counts, key=counts.get)
    if counts[best] == 0:
        return None, counts
    return best, counts


def split_header(header_line: str, delim: str) -> list[str]:
    """
    Разбиваем строку заголовков на колонки.
    """
    return [h.strip().replace("\ufeff", "") for h in header_line.rstrip().split(delim)]


def parse_rows_iter(text: str, delim: str) -> tuple[Iterator[list[str]], list[str]]:
    """
    Потоковый разбор CSV:
    - header: список колонок
    - rows_iter: итератор по строкам данных (без list(reader))
    """
    f = io.StringIO(text)
    reader = csv.reader(f, delimiter=delim)

    header_row = next(reader, [])
    header = [c.strip() for c in header_row] if header_row else []

    return reader, header


# -------------------------------------------------
# Сравнение и нормализация заголовков
# -------------------------------------------------


def norm(s: str) -> str:
    """
    Нормализация строки для мягкого сравнения:
    без BOM, без лишних пробелов, без учёта регистра.
    """
    s = s.replace("\ufeff", "").strip()
    s = re.sub(r"\s+", " ", s)
    return s.casefold()


def compare_headers_strict(actual: list[str], expected: list[str]):
    """
    Строгое сравнение заголовков по позициям.
    """
    mismatches = []
    for i in range(max(len(actual), len(expected))):
        exp = expected[i] if i < len(expected) else None
        act = actual[i] if i < len(actual) else None
        if exp != act:
            mismatches.append((i + 1, exp, act))
    return mismatches


def build_lenient_mapping(actual: list[str], expected: list[str]):
    """
    Пытаемся безопасно сопоставить заголовки,
    если отличаются только регистр / пробелы.
    """
    if len(actual) != len(expected):
        return None, ["Количество колонок не совпадает"]

    mapping = {}
    used = set()
    problems = []

    exp_map = {norm(e): e for e in expected}

    for a in actual:
        key = norm(a)
        if key not in exp_map:
            problems.append(f"Неизвестная колонка: '{a}'")
            continue
        e = exp_map[key]
        if e in used:
            problems.append(f"Дублирующая колонка: '{a}' -> '{e}'")
            continue
        mapping[a] = e
        used.add(e)

    if problems or len(used) != len(expected):
        return None, problems

    return mapping, []


def normalized_header_for_fixed(
    actual_header: list[str], expected: list[str]
) -> tuple[list[str], list[str]]:
    """
    Возвращает:
      - fixed_header: заголовок ровно как expected (если удалось сопоставить)
      - problems: список проблем (если не удалось)
    """
    # Сначала попробуем "мягкое сопоставление"
    mapping, problems = build_lenient_mapping(actual_header[: len(expected)], expected)
    if problems:
        # Мягко не получилось - вернем строгие несовпадения как проблемы
        mism = compare_headers_strict(actual_header[: len(expected)], expected)
        problems2 = [
            f"Колонка {i}: ожидается '{exp}', получено '{act}'"
            for (i, exp, act) in mism
            if exp != act
        ]
        return expected[:], (problems + problems2)

    # Если mapping есть — значит различия только в регистре/пробелах/BOM
    # FIXED заголовок должен быть строго expected
    return expected[:], []


QUOTE_CHARS = "'\"`’‘“”«»"


def norm_mask(mask: str) -> str:
    m = (mask or "").replace("\ufeff", "").strip()
    # убираем внешние кавычки, апострофы
    m = m.strip(QUOTE_CHARS).strip()
    m = m.replace("/", "/").replace("/", "/").replace("⁄", "/")
    return m


def parse_mask_prefix(mask: str) -> int | None:
    m = norm_mask(mask)
    mm = MASK_RE.search(m)
    if not mm:
        return None
    try:
        return int(mm.group(1))
    except ValueError:
        return None


def parse_ip_from_ipport(value: str) -> str | None:
    s = (value or "").strip()
    if not s:
        return None

    # [IPv6]:port
    m = re.match(r"^\[([0-9a-fA-F:]+)\]:(\d{1,5})$", s)
    if m:
        return m.group(1)

    # IPv4:port
    m = re.match(r"^(\d{1,3}(?:\.\d{1,3}){3}):(\d{1,5})$", s)
    if m:
        return m.group(1)

    # IPv6:port
    # # 2a00....:443
    if ":" in s:
        left, sep, right = s.rpartition(":")
        if sep and right.isdigit():
            return left

    return None


def build_fixed_csv(
    csv_file: Path,
    text: str,
    delim: str,
    expected_cols: list[str],
    out_dir: Path,
) -> Path:
    """
    Делает временный CSV, который "съест" main.py:
    - заголовок приводим к expected_cols (по позиции)
    - строки обрезаем до 16 или дополняем пустыми до 16
    - лишние колонки от хвостовых ||| игнорируем
    - пишем UTF-8 (без BOM) c тем же разделителем '|'
    """

    out_dir.mkdir(parents=True, exist_ok=True)
    fixed_path = out_dir / f"{csv_file.stem}_FIXED.csv"

    lines = text.splitlines()
    if not lines:
        raise ValueError("Файл пустой")

    reader = csv.reader(lines, delimiter=delim)
    rows = list(reader)
    if not rows:
        raise ValueError("Не удалось прочитать CSV")

    header = rows[0]
    data_rows = rows[1:]

    # нормализуем заголовки и приводим к длине expected
    # header = [normalized_header_for_fixed(x, expected_cols) for x in header]
    # header16 = (header + [""] * len(expected_cols))[: len(expected_cols)]
    header16 = expected_cols[:]

    # ВАЖНО: чтобы main.py не падал на заголовках — делаем их ровно ожидаемыми
    header16 = expected_cols[:]  # именно так: фиксируем строго

    fixed_rows = [header16]

    for r in data_rows:
        # обрезаем/дополняем до 16
        r16 = (r + [""] * len(expected_cols))[: len(expected_cols)]
        fixed_rows.append(r16)

    # пишем как UTF-8 без BOM, newline = "" важно для csv в винде
    with open(fixed_path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter="|", quoting=csv.QUOTE_MINIMAL)
        w.writerows(fixed_rows)

    return fixed_path


# -------------------------------------------------
# Проверка заполнения и логики
# -------------------------------------------------


def is_empty(v) -> bool:
    return v is None or str(v).strip() == ""


def contains_token(method: str, token: str) -> bool:
    parts = re.split(r"[,\s;]+", (method or "").lower())
    return token.lower() in parts


def precheck_filling_and_logic(row: dict, rownum: int) -> list[str]:
    """
    Проверка пустых полей и логических зависимостей.
    """
    errors = []

    # Обязательные поля
    required = [
        "Number",
        "Название сервиса",
        "юр_лицо",
        "ОГРН",
        "IPv4/IPv6",
        "Сеть в формате network ID",
        "MASK",
        "CDN Yes/No",
        "Protocol",
        "Способ проверки",
    ]

    for field in required:
        if is_empty(row.get(field)):
            errors.append(f"Строка {rownum}: обязательное поле пустое:{field}")
    method = row.get("Способ проверки", "")

    if contains_token(method, "curl") and is_empty(row.get("URL проверки")):
        errors.append(f"Строка {rownum}: указан curl, но URL проверки не указан")

    if contains_token(method, "telnet") and is_empty(row.get("Проверочный IP+Порт")):
        errors.append(
            f"Строка {rownum}: указан telnet, но Проверочный IP+Порт не указан"
        )

    if contains_token(method, "dig") and is_empty(row.get("DNS Host")):
        errors.append(f"Строка {rownum}: указан dig, но DNS Host не указан")

    if row.get("Protocol") and is_empty(row.get("Port")):
        errors.append(f"Строка {rownum}: указан Protocol, но Port не указан")

    return errors


def cell(row: list[str], idx: int) -> str:
    return (row[idx] if idx < len(row) else "").strip()


def split_csv_list(value: str) -> list[str]:
    """
    Разбивает список значений в одной ячейке.
    Поддерживаем: "a,b", "a; b", "a b" (на всякий случай).
    """
    v = (value or "").strip().lower()
    if not v:
        return []
    v = v.replace(";", ",")
    parts = [p.strip() for p in v.split(",") if p.strip()]
    return parts


def normalize_protocols(value: str) -> set[str]:
    """
    Protocol может быть: tcp, udp, оба, tcp,udp, udp,tcp, tcp/udp.
    Возвращаем множество {'tcp','udp'} и т.п.
    """

    # гарантируем, что v всегда определена
    v = (value or "").strip().lower()
    if not v:
        return set()

    v = v.replace("/", ",").replace(";", ",")
    v = re.sub(r"\s+", "", v)

    if v in {"оба", "both"}:
        return {"tcp", "udp"}

    parts = [p for p in v.split(",") if p]
    out: set[str] = set()

    for p in parts:
        if p in {"оба", "both"}:
            out.update({"tcp", "udp"})
        else:
            out.add(p)

    return out


_ipv4_re = re.compile(r"^\d{1,3}(\.\d{1,3}){3}$")
_bracket_ipv6_port_re = re.compile(r"^\[.+\]:(\d{1,5})$")
_ipv4_port_re = re.compile(r"^(\d{1,3}(\.\d{1,3}){3}):(\d{1,5})$")


def looks_like_ip_port(value: str) -> bool:
    v = (value or "").strip()
    if not v:
        return False

    m = _ipv4_port_re.match(v)
    if m:
        port = int(m.group(3))
        return 1 <= port <= 65535

    m2 = _bracket_ipv6_port_re.match(v)
    if m2:
        port = int(m2.group(1))
        return 1 <= port <= 65535

    # для ping допускаем просто IPv4 (без порта)
    if _ipv4_re.match(v):
        return True

    return False


def validate_row_rules(row: list[str]) -> list[str]:
    """
    Возвращает список ошибок по строке.
    Проверки основаны на ТЗ:
    - пустые обязательные поля
    - Protocol ↔ Port
    - Способ проверки ↔ URL ↔ Проверочный IP+Порт
    """
    errs: list[str] = []

    # Индексы по ожидаемым колонкам (фиксированные позиции)
    IDX_PORT = 7
    IDX_PROTOCOL = 9
    IDX_DNS = 10
    IDX_SNI = 11
    IDX_CHECK_IPPORT = 12
    IDX_URL = 13
    IDX_METHOD = 14
    IDX_NETWORK = 5
    IDX_MASK = 6

    port = cell(row, IDX_PORT)
    protocol_raw = cell(row, IDX_PROTOCOL)
    dns = cell(row, IDX_DNS)
    sni = cell(row, IDX_SNI)
    check_ipport = cell(row, IDX_CHECK_IPPORT)
    url = cell(row, IDX_URL)
    method_raw = cell(row, IDX_METHOD)
    network_id = cell(row, IDX_NETWORK)
    mask = cell(row, IDX_MASK)

    # 1) Пустые обязательные ячейки (по REQUIRED_COLUMNS из config.colomns)
    for col in REQUIRED_COLUMNS:
        v = cell(row, col.index)
        if is_empty(v):
            errs.append(f"Пустое обязательное поле: '{col.name}'")

    # 2) Protocol: tcp/udp/оба
    protocols = normalize_protocols(protocol_raw)
    if protocols:
        bad = [p for p in protocols if p not in VALID_PROTOCOLS]
        if bad:
            errs.append(
                f"Недопустимое значение Protocol: {bad} (ожидается tcp/udp или оба)"
            )
    # если указан Protocol, но ячейка Port пустая - то это ошибка (собираем противоречие)
    if (not is_empty(protocol_raw)) and is_empty(port):
        errs.append("Указан Protocol, но ячейка Port пустая")
    if (not is_empty(port)) and is_empty(protocol_raw):
        errs.append("Указан Port, но ячейка Protocol пустая")

    # 3) Способ проверки: может быть список
    methods = [m.strip().lower() for m in split_csv_list(method_raw)]

    # (если поле пустое — оно уже попадёт как "пустое обязательное", если в REQUIRED_COLUMNS)
    for m in methods:
        if m not in VALID_CHECK_METHODS:
            errs.append(f"Недопустимый способ проверки: '{m}'")

    # 4) Правила по методам из таблицы Word

    if "telnet" in methods:
        # - telnet: обязателен Проверочный IP+Порт; URL обычно пуст
        if is_empty(check_ipport):
            # строгий режим: порт в IP+Port должен совпадать с колонкой Port
            if is_empty(check_ipport):
                errs.append("Указан telnet, но ячейка 'Проверочный IP+Порт' пустая")
            # URL не нужен, но не считаем ошибкой, если заполнен

    if "ping" in methods:
        # - ping: нужен IP (берем из Проверочный IP+Порт)
        if is_empty(check_ipport) and is_empty(dns):
            errs.append("Указан ping, но нет ни IP, ни DNS для проверки")

    if "dig" in methods:
        # - dig: нужен DNS Host (или SNI/HTTP HOST), URL не нужен
        if is_empty(dns) and is_empty(sni):
            errs.append(
                "Указан dig, но ячейки 'DNS Host' и 'SNI/HTTP HOST' пустые (нужно имя для проверки)"
            )

    # 5) HTTPS + TCP => SNI/HTTP HOST обязателен
    if url.lower().startswith("https://") and ("tcp" in protocols):
        if is_empty(sni):
            errs.append("Для HTTPS (tcp) обязательно поле 'SNI/HTTP HOST'")

    # 6) Формат Проверочного IP+Порт (строго)
    parsed_ip, parsed_port = (
        _parse_ip_port(check_ipport) if not is_empty(check_ipport) else (None, None)
    )
    if not is_empty(check_ipport) and parsed_ip is None:
        errs.append(
            "Проверочный IP+Порт должен быть в формате IPv4:Port, например 1.2.3.4: 443"
        )

    # 6.1) Проверка: IP из "Проверочный IP+Порт" принадлежит сети network_id/mask

    # MASK
    # if not is_empty(mask):
    #    p = parse_mask_prefix(mask)
    #    if p is None:
    #        errs.append(f"MASK: неверный формат (ожидается '/число'): '{mask!r}'")

    # NETWORK
    if not is_empty(network_id) and not is_empty(mask) and not is_empty(check_ipport):
        ip_str = parse_ip_from_ipport(check_ipport)
        if ip_str is None:
            errs.append(
                f"Проверочный IP+Порт: неверный формат (ожидается ip:port): '{check_ipport}'"
            )
        else:
            prefix = parse_mask_prefix(mask)
            if prefix is None:
                errs.append(f"MASK: неверный формат (ожидается '/число'): '{mask!r}'")
            else:
                try:
                    ip_obj = ipaddress.ip_address(ip_str)
                    net_obj = ipaddress.ip_network(
                        f"{network_id}/{prefix}", strict=False
                    )

                    # контроль версии: v4/v6 должны совпадать
                    if ip_obj.version != net_obj.version:
                        errs.append(
                            f"Несовпадение IPv4/IPv6: сеть '{network_id}{mask}', проверочный IP '{ip_str}'"
                        )
                    elif ip_obj not in net_obj:
                        errs.append(
                            f"Проверочный IP '{ip_str}' не принадлежит сети '{net_obj}'"
                        )
                except ValueError:
                    errs.append(
                        f"Сеть.MASK: некорректные значения: network='{network_id}', mask='{mask}'"
                    )

    # 7) URL проверки: строго по форме
    # curl -> URL обязателен и только http(s)
    # telnet/ping/dig -> URL должен быть пустым( чтобы не было мусора)

    if "curl" in methods:
        # для curl URL обязателен и должен быть http(s)
        if is_empty(url):
            errs.append("Указан curl, но ячейка 'URL проверки' пустая")
        elif not _is_http_url(url):
            errs.append(
                f"URL проверки при curl должен начинаться с http:// или https:// (получено: '{url}')"
            )
        if is_empty(url):
            errs.append(
                "URL проверки должна быть пустая для метода(ов) {methods} (получено: '{url}')"
            )
    else:
        # для всех остальных методов URL НЕ обязателен
        # но если заполнен - просто проверим, что это похоже на URL
        if not is_empty(url):
            if not (
                url.lower().startswith("http://")
                or url.lower().startswith("https://")
                or "." in url  # разрешаем домены типа v-a.yandex.ru
            ):

                errs.append(f"URL проверки имеет формат: '{url}')")
    return errs


# -------------------------------------------------
# Запуск main.py validate
# -------------------------------------------------


def run_main_validate(
    project_dir: Path,
    fixed_csv: Path,
    out_csv: Path,
    skip_whois: bool = False,
) -> tuple[str, int]:
    """
    Запускает main.py validate и возвращает (stdout + stderr, returncode)
    При этом печатает вывод main в консоль вживую
    """
    cmd = [
        sys.executable,
        str(project_dir / "main.py"),
        "validate",
        "--input", str(fixed_csv),
        "--output", str(out_csv),
        "--max-errors", "10000",
    ]
    if skip_whois:
        cmd.append("--skip-whois")

    env = dict(os.environ)
    env["PYTHONUTF8"] = "1"
    env["PYTHONIOENCODING"] = "utf-8"
    env["PYTHONUNBUFFERED"] = "1"

    p = subprocess.Popen(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="replace",
        env=env,
        cwd=str(project_dir),  # важно: main.py любит запуск из своей папки
        bufsize=1,
    )

    out_lines = []
    assert p.stdout is not None
    for line in p.stdout:
        out_lines.append(line)
        print(line, end="", flush=True)

    rc = p.wait()    
    return "".join(out_lines), rc

def write_text(path: Path, text: str):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8-sig")


def format_header_mismatches(mism) -> str:
    lines = []
    for item in mism:
        try:
            col_no, expected, got = item
            lines.append(f"Колонка {col_no}: ожидается *{expected}*, получено *{got}*")
        except Exception:
            lines.append(str(item))
    return "\n".join(lines)


IPV4_RE = re.compile(r"^(?:\d{1,3}\.){3}\d{1,3}$")
IPPORT_RE = re.compile(r"^(.+):(\d{1,5})$")


def _is_valid_ipv4(ip: str) -> bool:
    ip = (ip or "").strip()
    if not IPV4_RE.match(ip):
        return False
    parts = ip.split(".")
    return all(0 <= int(p) <= 255 for p in parts)


def _parse_ip_port(value: str):
    """Возвращает (ip, port_int) или (None,  None) если формат неверный"""
    s = (value or "").strip()
    m = IPPORT_RE.match(s)
    if not m:
        return None, None

    ip = m.group(1).strip()
    port_str = m.group(2).strip()

    try:
        port = int(port_str)
    except ValueError:
        return None, None

    if port < 1 or port > 65535:
        return None, None

    if not _is_valid_ipv4(ip):
        return None, None

    return ip, port


def _is_http_url(url: str) -> bool:
    s = (url or "").strip().lower()
    return s.startswith("http://") or s.startswith("https://")


@dataclass
class FileOutcome:
    file: Path
    ran_validator: bool


MAX_EXAMPLES_PER_TYPE = 10000


def add_group(counts: dict, rows_all: dict, key: str, line_no: int | None = None):
    """Увеличивает счетчик ошибок и сохраняет номера строк-примеров"""
    counts[key] += 1
    if line_no is not None:
        rows_all[key].append(line_no)


def dump_groups(rs, title: str, counts: dict, rows_all: dict):
    """Печатает сгруппированные ошибки по убыванию частоты"""
    if not counts:
        rs.print_startup_warning(f"{title}: нет")
        return

    rs.print_startup_warning(title + ":")

    # сортируем по количеству (самые частые сверху)
    for msg, cnt in sorted(counts.items(), key=lambda kv: kv[1], reverse=True):
        all_rows = sorted(set(rows_all.get(msg, [])))
        if all_rows:
            rs.print_startup_warning(
                f"• ({cnt} шт.) {msg}\n Строки: {compress_ranges(all_rows)}"
            )
        else:
            rs.print_startup_warning(f"• ({cnt} шт.) {msg}")


def build_excel_report(
    main_csv: Path,
    excel_path: Path,
    sep: str,
    precheck_txt: Path | None = None,
    main_log_txt: Path | None = None,
) -> None:
    """
    Делает Excel-отчёт:
      - MAIN_RESULT: таблица из main.csv
      - MAIN_LOG: лог main.py (если есть)
      - PRECHECK: precheck-отчёт (если есть)
    """
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    # 1) Таблица main.py -> Excel
    df = pd.read_csv(main_csv, sep=sep, encoding="utf-8")
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="MAIN_RESULT", index=False)

    # 2) Добавляем текстовые листы
    wb = load_workbook(excel_path)

    def add_text_sheet(sheet_name: str, txt_path: Path):
        if not txt_path or not txt_path.exists():
            return
        ws = wb.create_sheet(sheet_name)
        text = txt_path.read_text(encoding="utf-8", errors="replace")
        # пишем построчнов колонку А\
        for i, line in enumerate(text.splitlines(), start=1):
            ws.cell(row=i, column=1, value=line)

    add_text_sheet("MAIN_LOG", main_log_txt)
    add_text_sheet("PRECHECK", precheck_txt)

    # Удобно: закрепим шапку таблицы
    ws0 = wb["MAIN_RESULT"]
    ws0.freeze_panes = "A2"
    ws0.auto_filter.ref = ws0.dimensions

    wb.save(excel_path)


def compress_ranges(nums: list[int]) -> str:
    """Например [2,3,4,7,8,10] - '2-4', '7-8', 10"""
    if not nums:
        return ""

    nums = sorted(set(nums))
    ranges = []
    start = prev = nums[0]

    for x in nums[1:]:
        if x == prev + 1:
            prev = x
        else:
            ranges.append((start, prev))
            start = prev = x

    ranges.append((start, prev))

    parts = []
    for a, b in ranges:
        parts.append(f"{a}-{b}" if a != b else f"{a}")

    return ", ".join(parts)


def _clean_main_log_for_report(text: str) -> str:
    """
    Чистим main.txt для общего отчета:
    - убираем tqdm/прогресс бары
    - выкидываем REQUIRED (они уже есть в precheck)
    """
    if not text:
        return ""

    out_lines = []
    for line in text.splitlines():
        s = line.rstrip("\n")

        # 1) выкидываем прогресс бары и мусор
        if s.startswith("Валидация:"):
            continue
        if "Whois/RDAP запросы" in s:
            continue
        if re.search(r"\|\s*\d+/\d+\s*\[", s):  # что то вроде | 9/104 [00:06...]
            continue
        if s in ("[A", ""):
            continue

        # 2) выкидываем REQUIRED из main
        if "[REQUIRED]" in s:
            continue

        out_lines.append(line)

        # дополнительно: схлопнем много пустых строк подряд до одной
    cleaned = "\n".join(out_lines)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned).strip()
    return cleaned


def build_final_report(precheck_path: Path, main_path: Path, final_path: Path) -> None:
    """
    Склеиваем 2 отчета в 1 общий
    - precheck кладем целиком
    - main чистим от REQUIRED и от прогресс бара
    """
    pre_text = (
        precheck_path.read_text(encoding="utf-8", errors="replace")
        if precheck_path.exists()
        else ""
    )
    main_text_raw = (
        main_path.read_text(encoding="utf-8", errors="replace")
        if main_path.exists()
        else ""
    )
    main_text = _clean_main_log_for_report(main_text_raw)

    final = []
    final.append("==== PRECHECK ====\n")
    final.append(pre_text.strip() + "\n")

    final.append("\n==== MAIN ====\n")
    final.append(
        (main_text.strip() + "\n")
        if main_text.strip()
        else "(main.txt не найден или пуст)\n"
    )

    final_path.write_text("".join(final), encoding="utf-8")


def parse_main_errors(main_text: str) -> list[tuple[str, str]]:
    """
    Парсим ошибки из MAIN.txt.
    Возвращаем список пар (код_ошибки, текст).
    Пример:
      ("REQUIRED", "Пустое обязательное поле: 'CDN Yes/No'")
    """
    pairs = []

    for line in main_text.splitlines():
        line = line.strip()
        if not line.startswith("[❌]"):
            continue

        # [❌] Number 123: [REQUIRED] CDN Yes/No: Обязательное поле не заполнено
        m = re.search(r"\[(\w+)\]\s+(.*)", line)
        if not m:
            continue

        code = m.group(1)
        text = m.group(2)
        pairs.append((code, text))

    return pairs


def group_main_errors(pairs: list[tuple[str, str]]) -> dict[str, list[str]]:
    """
    Группируем ошибки MAIN по тексту.
    Возвращает:
      {
        "CDN Yes/No: Обязательное поле не заполнено": ["REQUIRED", "REQUIRED", ...]
      }
    """
    groups = defaultdict(list)

    for code, text in pairs:
        groups[text].append(code)

    return groups


def format_grouped_main(groups: dict[str, list[str]]) -> str:
    """
    Форматируем сгруппированные ошибки MAIN для отчёта.
    """
    if not groups:
        return "Ошибок MAIN не обнаружено. \n"

    lines = []
    for text, codes in groups.items():
        count = len(codes)
        lines.append(f"• ({count} шт.) {text}")

    return "\n".join(lines)


# -------------------------------------------------
# ЕДИНАЯ ТОЧКА ВХОДА
# -------------------------------------------------


def main():
    print("=== MAIN ENTERED ===", flush=True)
    parser = argparse.ArgumentParser(
        description="Precheck CSV + сбор всех ошибок в отчет"
    )
    parser.add_argument("--input", "-i", required=True, help="CSV файл или папка с CSV")
    args = parser.parse_args()

    inp = Path(args.input)

    # Защита: проверяем, что это CSV или папка с CSV
    if inp.is_file():
        if inp.suffix.lower() != ".csv":
            print("ОШИБКА: Входной файл не является CSV:", inp.name)
            print("Пожалуйста, перетащите файл с расширением .csv")
            return
    else:
        csv_file = list(inp.glob("*.csv"))
        if not csv_file:
            print("ОШИБКА: в папке нет ни одного CSV файла:", inp)
            return

    files = [inp] if inp.is_file() else csv_file

    reports_all_dir = BASE_DIR / "reports_all"
    reports_all_dir.mkdir(parents=True, exist_ok=True)

    report_only_dir = BASE_DIR / "report_only"
    report_only_dir.mkdir(parents=True, exist_ok=True)

    for csv_file in files:

        # Группировка всех проблем
        counts = defaultdict(int)  # текст проблемы -> сколько раз
        rows_all = defaultdict(list)  # ВСЕ номера строк

        precheck_report_path = reports_all_dir / f"{csv_file.stem}_PRECHECK.txt"

        # 1) читаем файл как текст
        print("1) before detect_encoding", flush=True)
        enc, text = detect_encoding_and_text(csv_file)
        print(f"2) after detect_encoding: {enc}, bytes={len(text)} chars", flush=True)

        # 2) открываем отчет и записываем грамотно
        with open(precheck_report_path, "w", encoding="utf-8-sig") as f:
            rs = ReportService(output=f)

            rs.print_startup_warning(f"Проверка файла: {csv_file.name}")

            enc_norm = enc.lower()

            if enc_norm.startswith("utf-8"):
                rs.print_startup_warning("Кодировка: UTF-8")
            else:
                rs.print_startup_warning(f"Кодировка: {enc}")
                rs.print_startup_warning("ВНИМАНИЕ: файл CSV сохранен не в UTF-8. ")

            lines = text.splitlines()
            print(f"3) after splitlines: lines={len(lines)}", flush=True)
            if not lines:
                rs.print_error("Файл пустой")
                continue

            header_line = lines[0] if lines else ""
            print(f"4) header sample: {header_line[:80]!r}", flush=True)

            # по ТЗ разделитель строго |

            if "|" in header_line:
                delim = "|"
            elif "¦" in header_line:
                delim = "¦"
                rs.print_startup_warning(
                    "Обнаружен нестандартный разделитель '¦' (Excel)."
                )
            elif "│" in header_line:
                delim = "│"
                rs.print_startup_warning("Обнаружен нестандартный разделитель '│'.")
            else:
                rs.print_error(
                    "Не найден разделитель '|'. Файл должен быть сохранен как CSV с разделителем |"
                )
                delim = "|"

            # 3) Парсим CSV
            rows, header = parse_rows_iter(text, delim)
            print(f"5) after parse_rows_iter: header_cols={len(header)}", flush=True)

            # Сколько колонок реально в файле
            rs.print_startup_warning(
                f"Колонок в файле: {len(header)} (по Инструкции должно быть {len(COLUMN_NAMES)})"
            )

            mism = compare_headers_strict(header[: len(COLUMN_NAMES)], COLUMN_NAMES)
            if mism:
                # одну запись "тип проблемы" + детали отдельными пунктами
                rs.print_startup_warning(
                    "Заголовки: есть расхождения (регистр/написание):"
                )
                for col_no, expected, got in mism:
                    rs.print_startup_warning(
                        f"  - Колонка {col_no}: '{got}' -> должно быть '{expected}'"
                    )
            else:
                rs.print_startup_warning("Заголовки: ОК")

            # 5) Проверяем строки и собираем ВСЕ ошибки

            first_row = next(rows, None)
            if first_row is None:
                rs.print_error("В файле нет строк данных (только заголовок)")
                continue

            # вернем первую строку обратно в поток
            rows = itertools.chain([first_row], rows)

            total = 0
            bad = 0

            extra_cols_noted = False

            extra_rows = []  # номера строк, где есть лишние колонки
            extra_by_amount = {}  # опционально: сколько лишних колонок  -> списки строк

            print("6) before row loop", flush=True)

            for i, row in enumerate(rows, start=2):  # start=2 потому что 1- заголовок
                # пропускаем полностью пустые строки
                if all((c or "").strip() == "" for c in row):
                    continue

                total += 1
                row_has_problem = False

                # 1) Проверка количества колонок
                if len(row) < len(COLUMN_NAMES):
                    row_has_problem = True
                    rs.print_error(
                        f"Строка {i}: количество колонок  = {len(row)}, ожидается {len(COLUMN_NAMES)} "
                        "(не хватает колонок)"
                    )
                    add_group(
                        counts,
                        rows_all,
                        f"Строки: не хватает колонок (меньше {len(COLUMN_NAMES)})",
                        i,
                    )
                    row16 = row + [""] * (len(COLUMN_NAMES) - len(row))

                elif len(row) > len(COLUMN_NAMES):
                    extra = len(row) - len(COLUMN_NAMES)

                    extra_rows.append(i)

                    # опционально: группировка по величине лишних колонок
                    extra_by_amount.setdefault(extra, []).append(i)

                    row16 = row[: len(COLUMN_NAMES)]
                else:
                    row16 = row

                # 2) логические правила заполнения
                errs = validate_row_rules(row16)
                if errs:
                    row_has_problem = True
                    for e in errs:
                        add_group(counts, rows_all, e, i)

                # 3) финально считаем строку "битой"
                if row_has_problem:
                    bad += 1

            if extra_rows:
                rs.print_startup_warning("Лишние колонки (лишние '|') обнаружены:")
                for extra, rows_list in sorted(extra_by_amount.items()):
                    ranges = compress_ranges(rows_list)
                    rs.print_startup_warning(
                        f"   +{extra}: строки {ranges} (всего {len(rows_list)})"
                    )

            print("7) after row loop", flush=True)        

            rs.print_startup_warning(f"Итого строк данных: {total}")
            # rs.print_startup_warning(f"Строк с ошибками: {bad}")

            # 6) Готовим FIXED CSV и запускаем main.py

           

            fixed_dir = reports_all_dir / "_temp"
            main_out_csv = reports_all_dir / f"{csv_file.stem}_MAIN_result.csv"
            main_log_txt = reports_all_dir / f"{csv_file.stem}_MAIN.txt"

            main_log = ""
            rc = 999  # любое неуспешно по умолчанию

            try:
                print("8) before build_fixed_csv", flush=True)
                fixed_csv = build_fixed_csv(
                    csv_file=csv_file,
                    text=text,
                    delim=delim,  # тот, который мы определили
                    expected_cols=COLUMN_NAMES,  # 16 наших ожидаемых
                    out_dir=fixed_dir,
                )

                print("9) after build_fixed_csv", flush=True)

                print("10) before run_main_validate", flush=True)

                main_log, rc = run_main_validate(
                    project_dir=Path(__file__).parent,  # папка, где лежит main.py
                    fixed_csv=fixed_csv,
                    out_csv=main_out_csv,
                    skip_whois=False,
                )
                print(f"11) after run_main_validate rc={rc}", flush=True)

            except Exception as e:
                main_log = f"EXCEPTION: {e}"
                rc = 1
                rs.print_error(
                    f"Не удалось подготовить FIXED CSV / запустить main.py: {e}"
                )

                                  
            # ВСЕГДА пишем лог main.py
            main_log_txt.write_text(main_log, encoding="utf-8", errors="replace")

            # Итоговая сводка по PRECHECK
            dump_groups(rs, "Сводка проблем (сгруппировано)", counts, rows_all)

        print("Отчет сохранен:", precheck_report_path)

        # ✅ ВАЖНО: общий отчёт делаем ЗДЕСЬ, ПОСЛЕ try/except — всегда

        final_report_path = report_only_dir / f"{csv_file.stem}_REPORT.txt"
        build_final_report(precheck_report_path, main_log_txt, final_report_path)
        print(f"Общий отчет сохранен:", final_report_path.name)

    print("Предварительная проверка завершена")


if __name__ == "__main__":
    main()
